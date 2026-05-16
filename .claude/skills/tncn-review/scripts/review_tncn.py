"""
Skill 3 - tncn-review
Phan tich nhanh bang luong, xuat file tong quan thue TNCN (.xlsx).
"""
import sys, io
from collections import defaultdict
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

GIAM_TRU_BAN_THAN = 15_500_000
OUTPUT_DIR = Path("output")

# ── Styles ────────────────────────────────────────────────────────────────────
FONT_BASE    = Font(name="Arial", size=10)
FONT_BOLD    = Font(name="Arial", size=10, bold=True)
FONT_TITLE   = Font(name="Arial", size=11, bold=True)
FONT_SECTION = Font(name="Arial", size=10, bold=True, color="FFFFFF")

FILL_SECTION = PatternFill("solid", fgColor="2F5496")   # xanh dam
FILL_HEADER  = PatternFill("solid", fgColor="BDD7EE")   # xanh nhat
FILL_TOTAL   = PatternFill("solid", fgColor="D9D9D9")   # xam nhat
FILL_WARN    = PatternFill("solid", fgColor="FFE699")   # vang nhat

ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")

NUM_FMT = '#,##0'

thin = Side(style="thin", color="BFBFBF")
BORDER_THIN = Border(left=thin, right=thin, top=thin, bottom=thin)

def sc(ws, row, col, value, font=None, fill=None, align=None, num_fmt=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:    cell.font    = font
    if fill:    cell.fill    = fill
    if align:   cell.alignment = align
    if num_fmt: cell.number_format = num_fmt
    if border:  cell.border  = border
    return cell

# ── Tính thuế ─────────────────────────────────────────────────────────────────
def parse_vnd(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return int(round(v))
    try: return int(str(v).strip().replace(".","").replace(",","").replace(" ","") or 0)
    except: return 0

def calc_thue(tnct):
    if tnct <= 0: return 0
    t = min(tnct, 10_000_000) * 0.05
    if tnct > 10_000_000:  t += (min(tnct, 30_000_000)  - 10_000_000)  * 0.10
    if tnct > 30_000_000:  t += (min(tnct, 60_000_000)  - 30_000_000)  * 0.20
    if tnct > 60_000_000:  t += (min(tnct, 100_000_000) - 60_000_000)  * 0.30
    if tnct > 100_000_000: t += (tnct - 100_000_000) * 0.35
    return int(t)

# ── Đọc dữ liệu ───────────────────────────────────────────────────────────────
def read_data(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_rows = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        for i, row in enumerate(rows):
            cells = " ".join(str(c or "").lower() for c in row)
            if ("họ tên" in cells or "ho ten" in cells) and (
                "thu nhập" in cells or "thu nhap" in cells or "số tiền" in cells):
                header_idx = i; break
        if header_idx is None: continue
        headers = [str(c or "").strip() for c in rows[header_idx]]
        for row in rows[header_idx+1:]:
            if not any(c for c in row if c not in (None, "", 0)): continue
            rec = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
            rec["_sheet"] = sname
            all_rows.append(rec)
    return all_rows

# ── Tổng hợp ──────────────────────────────────────────────────────────────────
def aggregate(all_rows):
    def gv(rec, *keys):
        for k in keys:
            for hk in rec:
                if k.lower() in hk.lower():
                    v = parse_vnd(rec[hk])
                    if v: return v
        return 0

    def gs(rec, *keys):
        for k in keys:
            for hk in rec:
                if k.lower() in hk.lower():
                    v = str(rec[hk] or "").strip()
                    if v and v not in ("0", "None"): return v
        return ""

    periods = defaultdict(lambda: {
        "tn":0,"thue_kt":0,"bhxh":0,"cd":0,"gtgc":0,"has_bhxh":False
    })
    for rec in all_rows:
        name = str(next((rec[h] for h in rec if "họ" in h.lower()
                         or "ho ten" in h.lower()), "") or "").strip()
        if not name: continue
        period = gs(rec,"tháng trả lương","thang tra luong","tháng","thang") or "default"
        key = (name, period)
        bhxh = gv(rec,"BHXH khấu trừ","bhxh")
        gtgc = gv(rec,"Giảm trừ gia cảnh","giam tru gia canh","giam tru")
        periods[key]["tn"]      += gv(rec,"Tổng thu nhập","tong thu nhap","Số tiền","so tien")
        periods[key]["thue_kt"] += gv(rec,"Thuế TNCN đã khấu trừ","thue tncn da khau tru","Thuế TNCN","thue tncn")
        periods[key]["bhxh"]    += bhxh
        periods[key]["cd"]      += gv(rec,"Công đoàn","cong doan")
        periods[key]["gtgc"]    += gtgc
        if bhxh > 0: periods[key]["has_bhxh"] = True

    nv = defaultdict(lambda: {
        "tn":0,"thue_kt":0,"bhxh":0,"cd":0,"gtgc":0,"has_bhxh":False,"n_ky":0
    })
    for (name, period), d in periods.items():
        nv[name]["tn"]      += d["tn"]
        nv[name]["thue_kt"] += d["thue_kt"]
        nv[name]["bhxh"]    += d["bhxh"]
        nv[name]["cd"]      += d["cd"]
        nv[name]["gtgc"]    += d["gtgc"]
        if d["has_bhxh"]: nv[name]["has_bhxh"] = True
        if period != "default": nv[name]["n_ky"] += 1
    return nv

# ── Xuất xlsx ─────────────────────────────────────────────────────────────────
def write_xlsx(nv, filepath, out_path):
    pl01 = {n:d for n,d in nv.items() if d["has_bhxh"]}
    pl02 = {n:d for n,d in nv.items() if not d["has_bhxh"]}

    results = {}
    for name, d in nv.items():
        tnct    = max(0, d["tn"] - d["bhxh"] - d["cd"] - d["gtgc"])
        pt      = calc_thue(tnct)
        cl      = pt - d["thue_kt"]
        n_ky    = max(d["n_ky"], 1)
        tnct_tb = tnct / n_ky
        results[name] = {**d, "tnct": tnct, "thue_pt": pt, "chenh_lech": cl, "tnct_tb": tnct_tb}

    bac = [0]*5
    for r in results.values():
        tb = r["tnct_tb"]
        if tb <= 0:      bac[0] += 1
        elif tb <= 10e6: bac[1] += 1
        elif tb <= 30e6: bac[2] += 1
        elif tb <= 60e6: bac[3] += 1
        else:            bac[4] += 1

    t_tn    = sum(d["tn"]      for d in nv.values())
    t_bhxh  = sum(d["bhxh"]    for d in nv.values())
    t_kt    = sum(d["thue_kt"] for d in nv.values())
    t_pt    = sum(r["thue_pt"] for r in results.values())
    cl_tong = t_pt - t_kt

    top3     = sorted(results.items(), key=lambda x: -x[1]["thue_pt"])[:3]
    bot3     = sorted(results.items(), key=lambda x:  x[1]["tn"])[:3]
    canh_bao = sorted(
        [(n,r) for n,r in results.items() if abs(r["chenh_lech"]) > 5_000_000],
        key=lambda x: -abs(x[1]["chenh_lech"])
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TNCN Review"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22

    r = 1

    # Tiêu đề
    ws.merge_cells(f"A{r}:C{r}")
    sc(ws, r, 1, "TỔNG QUAN QUYẾT TOÁN THUẾ TNCN", font=Font(name="Arial",size=13,bold=True),
       align=ALIGN_CENTER)
    r += 1
    ws.merge_cells(f"A{r}:C{r}")
    sc(ws, r, 1, f"Luật 109/2025/QH15  |  File: {filepath}  |  Ngày tạo: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
       font=FONT_BASE, align=ALIGN_CENTER)
    r += 2

    def section(title):
        nonlocal r
        ws.merge_cells(f"A{r}:C{r}")
        sc(ws, r, 1, title, font=FONT_SECTION, fill=FILL_SECTION, align=ALIGN_LEFT)
        ws.row_dimensions[r].height = 18
        r += 1

    def row2(label, value, is_num=False, bold=False, fill=None, warn=False):
        nonlocal r
        f = FONT_BOLD if bold else FONT_BASE
        sc(ws, r, 1, label, font=f, fill=fill, align=ALIGN_LEFT,  border=BORDER_THIN)
        cell = sc(ws, r, 2, value, font=f, fill=fill,
                  align=ALIGN_RIGHT if is_num else ALIGN_LEFT,
                  num_fmt=NUM_FMT if is_num else None, border=BORDER_THIN)
        if warn and isinstance(value, (int,float)) and value > 0:
            cell.fill = FILL_WARN
        sc(ws, r, 3, "", border=BORDER_THIN)
        r += 1

    def blank():
        nonlocal r; r += 1

    # A. Tổng quan nhân sự
    section("A.  TỔNG QUAN NHÂN SỰ")
    row2("Tổng nhân viên",            len(nv),    is_num=True)
    row2("Có BHXH  (Phụ lục 01)",    len(pl01),  is_num=True)
    row2("Không BHXH  (Phụ lục 02)", len(pl02),  is_num=True)
    blank()

    # B. Chỉ số tài chính
    section("B.  CHỈ SỐ TÀI CHÍNH  (VND)")
    row2("Tổng thu nhập đã chi trả",           t_tn,    is_num=True)
    row2("Tổng BHXH đã khấu trừ",             t_bhxh,  is_num=True)
    row2("Thuế TNCN đã khấu trừ tại nguồn",   t_kt,    is_num=True)
    row2("Thuế TNCN phải nộp (Luật 109)",      t_pt,    is_num=True)
    row2("Chênh lệch  (+ phải nộp / − hoàn)", cl_tong, is_num=True, bold=True,
         fill=FILL_TOTAL, warn=(cl_tong > 0))
    blank()

    # C. Phân bổ bậc thuế
    section("C.  PHÂN BỔ BẬC THUẾ  (theo TNCT trung bình/kỳ)")
    sc(ws, r, 1, "Bậc",            font=FONT_BOLD, fill=FILL_HEADER, align=ALIGN_CENTER, border=BORDER_THIN)
    sc(ws, r, 2, "Điều kiện",      font=FONT_BOLD, fill=FILL_HEADER, align=ALIGN_CENTER, border=BORDER_THIN)
    sc(ws, r, 3, "Số người",       font=FONT_BOLD, fill=FILL_HEADER, align=ALIGN_CENTER, border=BORDER_THIN)
    r += 1
    for bac_info in [
        ("Không nộp",  "TNCT ≤ 0",              bac[0]),
        ("Bậc 1",      "TNCT ≤ 10 triệu",       bac[1]),
        ("Bậc 2",      "TNCT 10 – 30 triệu",    bac[2]),
        ("Bậc 3",      "TNCT 30 – 60 triệu",    bac[3]),
        ("Bậc 4–5",    "TNCT > 60 triệu",       bac[4]),
    ]:
        sc(ws, r, 1, bac_info[0], font=FONT_BASE, align=ALIGN_CENTER, border=BORDER_THIN)
        sc(ws, r, 2, bac_info[1], font=FONT_BASE, align=ALIGN_LEFT,   border=BORDER_THIN)
        sc(ws, r, 3, bac_info[2], font=FONT_BASE, align=ALIGN_CENTER,
           num_fmt=NUM_FMT, border=BORDER_THIN)
        r += 1
    blank()

    # D. Top & Bottom
    section("D.  TOP & BOTTOM")
    sc(ws, r, 1, "Top 3 người đóng thuế nhiều nhất", font=FONT_BOLD, align=ALIGN_LEFT)
    sc(ws, r, 2, "Thuế phải nộp (VND)",              font=FONT_BOLD, fill=FILL_HEADER,
       align=ALIGN_CENTER, border=BORDER_THIN)
    r += 1
    for i, (n, res) in enumerate(top3):
        sc(ws, r, 1, f"{i+1}.  {n}", font=FONT_BASE, align=ALIGN_LEFT,  border=BORDER_THIN)
        sc(ws, r, 2, res["thue_pt"],  font=FONT_BASE, align=ALIGN_RIGHT,
           num_fmt=NUM_FMT, border=BORDER_THIN)
        r += 1
    blank()
    sc(ws, r, 1, "3 người có thu nhập thấp nhất", font=FONT_BOLD, align=ALIGN_LEFT)
    sc(ws, r, 2, "Tổng thu nhập (VND)",            font=FONT_BOLD, fill=FILL_HEADER,
       align=ALIGN_CENTER, border=BORDER_THIN)
    r += 1
    for i, (n, res) in enumerate(bot3):
        sc(ws, r, 1, f"{i+1}.  {n}", font=FONT_BASE, align=ALIGN_LEFT,  border=BORDER_THIN)
        sc(ws, r, 2, res["tn"],       font=FONT_BASE, align=ALIGN_RIGHT,
           num_fmt=NUM_FMT, border=BORDER_THIN)
        r += 1
    blank()

    # E. Cảnh báo bất thường
    section(f"E.  CẢNH BÁO BẤT THƯỜNG  (|chênh lệch| > 5.000.000 VND)  —  {len(canh_bao)} trường hợp")
    if canh_bao:
        for col, hdr in [(1,"Họ tên"),(2,"Thuế đã KT"),(3,"Chênh lệch")]:
            sc(ws, r, col, hdr, font=FONT_BOLD, fill=FILL_HEADER, align=ALIGN_CENTER, border=BORDER_THIN)
        r += 1
        for n, res in canh_bao:
            sc(ws, r, 1, n,                   font=FONT_BASE, align=ALIGN_LEFT,  border=BORDER_THIN)
            sc(ws, r, 2, res["thue_kt"],       font=FONT_BASE, align=ALIGN_RIGHT,
               num_fmt=NUM_FMT, border=BORDER_THIN)
            cl_cell = sc(ws, r, 3, res["chenh_lech"], font=FONT_BASE, align=ALIGN_RIGHT,
               num_fmt=NUM_FMT, border=BORDER_THIN)
            cl_cell.fill = FILL_WARN
            r += 1
    else:
        ws.merge_cells(f"A{r}:C{r}")
        sc(ws, r, 1, "Không có trường hợp bất thường.", font=FONT_BASE, align=ALIGN_LEFT)
        r += 1

    blank()
    ws.merge_cells(f"A{r}:C{r}")
    sc(ws, r, 1, "Để tạo bộ hồ sơ quyết toán đầy đủ: dùng /tncn-generator hoặc /tncn-sheets-sync",
       font=Font(name="Arial", size=10, italic=True, color="595959"), align=ALIGN_LEFT)

    # Freeze header rows
    ws.freeze_panes = "A3"

    wb.save(str(out_path))

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    filepath = sys.argv[1] if len(sys.argv) > 1 else "Form bảng thu nhập học Claude code.xlsx"
    all_rows = read_data(filepath)
    nv = aggregate(all_rows)
    OUTPUT_DIR.mkdir(exist_ok=True)
    out = OUTPUT_DIR / f"tncn_review_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    write_xlsx(nv, filepath, out)
    print(f"OK: {out}")
