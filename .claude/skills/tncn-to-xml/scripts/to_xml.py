#!/usr/bin/env python3
"""
Skill: tncn-to-xml
Chuyen doi du lieu bang luong → XML quyet toan TNCN (05/QTT-TNCN)
Dung cung logic aggregate voi tncn-generator, dam bao so lieu khop.
"""
import sys, io, os, xml.etree.ElementTree as ET
from datetime import datetime
from collections import defaultdict
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

try:
    import xlrd
except ImportError:
    print("ERROR: pip install xlrd")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ── Hang so thue Luat 109/2025/QH15 ──────────────────────────────────────────
GIAM_TRU_BAN_THAN  = 15_500_000   # /thang
GIAM_TRU_PHU_THUOC =  6_200_000   # /thang/nguoi
OUTPUT_DIR = Path("output")


def parse_vnd(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return int(round(v))
    try:
        return int(str(v).strip().replace(".", "").replace(",", "").replace(" ", "") or 0)
    except:
        return 0


def calc_thue(tnct):
    """Tinh thue luy tien 5 bac — nam (x12 so voi thang)"""
    if tnct <= 0:
        return 0
    # Quy doi sang nam: brackets nam = brackets thang * 12
    t  = min(tnct, 120_000_000)  * 0.05
    if tnct > 120_000_000: t += (min(tnct, 360_000_000)   - 120_000_000)  * 0.10
    if tnct > 360_000_000: t += (min(tnct, 720_000_000)   - 360_000_000)  * 0.20
    if tnct > 720_000_000: t += (min(tnct, 1_200_000_000) - 720_000_000)  * 0.30
    if tnct > 1_200_000_000: t += (tnct - 1_200_000_000) * 0.35
    return int(t)


# ── Doc file bang luong (cung logic voi generate_tncn.py) ────────────────────
def read_input(filepath):
    ext = Path(filepath).suffix.lower()
    all_rows = []
    INCOME_KEYS = {"họ tên", "họ và tên", "tổng thu nhập", "thuế tncn",
                   "ho ten", "tong thu nhap"}

    def process_sheet(rows_iter, sheet_name):
        rows = list(rows_iter)
        header_idx = None
        for i, row in enumerate(rows):
            cells = " ".join(str(c).lower().strip() if c else "" for c in row)
            if sum(1 for k in INCOME_KEYS if k in cells) >= 2:
                header_idx = i
                break
        if header_idx is None:
            return 0
        headers = [str(c).strip() if c else "" for c in rows[header_idx]]
        count = 0
        for row in rows[header_idx + 1:]:
            if not any(c for c in row if c not in (None, "", 0)):
                continue
            rec = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
            name = str(next((rec[h] for h in rec if any(
                k in h.lower() for k in ["họ tên", "họ và tên", "ho ten"]
            )), "") or "").strip()
            if not name:
                continue
            rec["_sheet"] = sheet_name
            all_rows.append(rec)
            count += 1
        return count

    if ext == ".xls":
        wb = xlrd.open_workbook(filepath)
        for sname in wb.sheet_names():
            ws = wb.sheet_by_name(sname)
            rows_iter = (
                [ws.cell_value(r, c) for c in range(ws.ncols)]
                for r in range(ws.nrows)
            )
            n = process_sheet(rows_iter, sname)
            if n:
                print(f"  Sheet '{sname}': {n} dong")
    elif ext in (".xlsx", ".xlsm") and openpyxl:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows_iter = ([c.value for c in row] for row in ws.iter_rows())
            n = process_sheet(rows_iter, sname)
            if n:
                print(f"  Sheet '{sname}': {n} dong")
    else:
        print(f"ERROR: Dinh dang khong ho tro hoac thieu openpyxl: {ext}")
        sys.exit(1)

    return all_rows


def aggregate(all_rows):
    def gv(rec, *keys):
        for k in keys:
            for hk in rec:
                if k.lower() in hk.lower():
                    v = parse_vnd(rec[hk])
                    if v:
                        return v
        return 0

    def gs(rec, *keys):
        for k in keys:
            for hk in rec:
                if k.lower() in hk.lower():
                    v = str(rec[hk] or "").strip()
                    if v and v not in ("0", "None"):
                        return v
        return ""

    periods = defaultdict(lambda: {
        "tn": 0, "thue_kt": 0, "bhxh": 0, "cd": 0, "gtgc": 0,
        "cccd": "", "mst": "", "has_bhxh": False,
    })

    for rec in all_rows:
        name = str(next((rec[h] for h in rec if any(
            k in h.lower() for k in ["họ tên", "họ và tên", "ho ten"]
        )), "") or "").strip()
        if not name:
            continue

        period = gs(rec, "tháng trả lương", "thang tra luong", "tháng", "thang") or "default"
        key = (name, period)

        bhxh = gv(rec, "BHXH khấu trừ", "bhxh khau tru", "bhxh")
        gtgc = gv(rec, "Giảm trừ gia cảnh", "giam tru gia canh", "giam tru")

        periods[key]["tn"]      += gv(rec, "Tổng thu nhập", "tong thu nhap", "Số tiền", "so tien")
        periods[key]["thue_kt"] += gv(rec, "Thuế TNCN đã khấu trừ", "thue tncn da khau tru", "Thuế TNCN", "thue tncn")
        periods[key]["bhxh"]    += bhxh
        periods[key]["cd"]      += gv(rec, "Công đoàn", "cong doan")
        periods[key]["gtgc"]    += gtgc
        if bhxh > 0:
            periods[key]["has_bhxh"] = True
        if not periods[key]["cccd"]:
            periods[key]["cccd"] = gs(rec, "cccd", "cmnd", "so cccd", "so dinh danh")
        if not periods[key]["mst"]:
            periods[key]["mst"] = gs(rec, "ma so thue", "mst")

    nv = defaultdict(lambda: {
        "tn": 0, "thue_kt": 0, "bhxh": 0, "cd": 0, "gtgc": 0,
        "cccd": "", "mst": "", "has_bhxh": False,
    })

    for (name, period), d in periods.items():
        nv[name]["tn"]      += d["tn"]
        nv[name]["thue_kt"] += d["thue_kt"]
        nv[name]["bhxh"]    += d["bhxh"]
        nv[name]["cd"]      += d["cd"]
        nv[name]["gtgc"]    += d["gtgc"]
        if d["has_bhxh"]:
            nv[name]["has_bhxh"] = True
        if not nv[name]["cccd"]:
            nv[name]["cccd"] = d["cccd"]
        if not nv[name]["mst"]:
            nv[name]["mst"] = d["mst"]

    return nv


# ── Xuat XML ──────────────────────────────────────────────────────────────────
def build_xml(nv, year):
    pl01 = {n: d for n, d in nv.items() if d["has_bhxh"]}
    pl02 = {n: d for n, d in nv.items() if not d["has_bhxh"]}

    t_tn = sum(d["tn"] for d in nv.values())
    t_kt = sum(d["thue_kt"] for d in nv.values())
    t_pt = sum(
        calc_thue(max(0, d["tn"] - d["bhxh"] - d["cd"] - d["gtgc"]))
        for d in nv.values()
    )

    root = ET.Element("HoSoQuyetToanTNCN")
    root.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
    root.set("phienBan", "1.0")

    # ── Thong tin chung ───────────────────────────────────────────────────────
    ttc = ET.SubElement(root, "ThongTinChung")
    ET.SubElement(ttc, "MauSo").text          = "05/QTT-TNCN"
    ET.SubElement(ttc, "TenMau").text         = "To khai quyet toan thue TNCN"
    ET.SubElement(ttc, "NamTinhThue").text    = str(year)
    ET.SubElement(ttc, "NgayXuatXML").text    = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    ET.SubElement(ttc, "CanCuPhapLy").text    = "Luat so 109/2025/QH15"
    ET.SubElement(ttc, "GiamTruBanThan").text = str(GIAM_TRU_BAN_THAN)
    ET.SubElement(ttc, "GiamTruPhuThuoc").text = str(GIAM_TRU_PHU_THUOC)

    # ── To khai chinh ─────────────────────────────────────────────────────────
    tk = ET.SubElement(root, "ToKhaiChinh")
    ET.SubElement(tk, "TongSoNLD").text           = str(len(nv))
    ET.SubElement(tk, "SoNLDCoHopDong").text      = str(len(pl01))
    ET.SubElement(tk, "SoNLDKhongHopDong").text   = str(len(pl02))
    ET.SubElement(tk, "TongThuNhapChiuThue").text = str(t_tn)
    ET.SubElement(tk, "TongThueKhauTru").text     = str(t_kt)
    ET.SubElement(tk, "TongThuePNop").text         = str(t_pt)
    ET.SubElement(tk, "ChenhLech").text            = str(t_pt - t_kt)

    # ── Phu luc 01 — Co hop dong lao dong (co BHXH) ──────────────────────────
    pl1_el = ET.SubElement(root, "PhuLuc01_CoHopDong")
    ET.SubElement(pl1_el, "TenPhuLuc").text = "05-1/BK-TNCN — NLD co hop dong lao dong"
    ET.SubElement(pl1_el, "SoNguoi").text   = str(len(pl01))

    for i, (name, d) in enumerate(sorted(pl01.items()), 1):
        tnct = max(0, d["tn"] - d["bhxh"] - d["cd"] - d["gtgc"])
        pt   = calc_thue(tnct)
        cl   = pt - d["thue_kt"]

        nnt = ET.SubElement(pl1_el, "NguoiNopThue")
        ET.SubElement(nnt, "STT").text            = str(i)
        ET.SubElement(nnt, "HoTen").text          = name
        ET.SubElement(nnt, "MST").text            = d["mst"]
        ET.SubElement(nnt, "CCCD").text           = d["cccd"]
        ET.SubElement(nnt, "TongThuNhap").text    = str(d["tn"])
        ET.SubElement(nnt, "BHXHKhauTru").text    = str(d["bhxh"])
        ET.SubElement(nnt, "GiamTruGiaCanh").text = str(d["gtgc"])
        ET.SubElement(nnt, "ThuNhapTinhThue").text = str(tnct)
        ET.SubElement(nnt, "ThueKhauTru").text    = str(d["thue_kt"])
        ET.SubElement(nnt, "ThuePNop").text       = str(pt)
        ET.SubElement(nnt, "ThueThua").text       = str(abs(cl) if cl < 0 else 0)
        ET.SubElement(nnt, "ThueCon").text        = str(cl if cl > 0 else 0)

    # ── Phu luc 02 — Khong hop dong lao dong (khong BHXH) ────────────────────
    pl2_el = ET.SubElement(root, "PhuLuc02_KhongHopDong")
    ET.SubElement(pl2_el, "TenPhuLuc").text = "05-2/BK-TNCN — CTV, khong hop dong lao dong"
    ET.SubElement(pl2_el, "SoNguoi").text   = str(len(pl02))

    for i, (name, d) in enumerate(sorted(pl02.items()), 1):
        nnt = ET.SubElement(pl2_el, "NguoiNopThue")
        ET.SubElement(nnt, "STT").text         = str(i)
        ET.SubElement(nnt, "HoTen").text       = name
        ET.SubElement(nnt, "MST").text         = d["mst"]
        ET.SubElement(nnt, "CCCD").text        = d["cccd"]
        ET.SubElement(nnt, "TongThuNhap").text = str(d["tn"])
        ET.SubElement(nnt, "ThueKhauTru").text = str(d["thue_kt"])

    return root


def indent_xml(elem, level=0):
    """Pretty-print XML (Python < 3.9 khong co ET.indent)"""
    indent = "\n" + "  " * level
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = indent + "  "
        if not elem.tail or not elem.tail.strip():
            elem.tail = indent
        for child in elem:
            indent_xml(child, level + 1)
        if not child.tail or not child.tail.strip():
            child.tail = indent
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = indent
    if not level:
        elem.tail = "\n"


def main():
    if len(sys.argv) < 2:
        print("Cach dung: python to_xml.py <file_bang_luong>")
        print("Vi du   : python to_xml.py output/bang_luong_from_sheets.xlsx")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    if not input_path.exists():
        print(f"ERROR: Khong tim thay '{input_path}'")
        sys.exit(1)

    year = datetime.now().year
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_path = OUTPUT_DIR / f"QuyetToanTNCN_{year}.xml"

    print(f"Doc du lieu : {input_path}")
    all_rows = read_input(str(input_path))
    if not all_rows:
        print("ERROR: Khong doc duoc du lieu tu file. Kiem tra lai cau truc sheet.")
        sys.exit(1)
    print(f"Tong        : {len(all_rows)} dong")

    print("Tong hop nhan su...")
    nv   = aggregate(all_rows)
    pl01 = {n: d for n, d in nv.items() if d["has_bhxh"]}
    pl02 = {n: d for n, d in nv.items() if not d["has_bhxh"]}
    print(f"  {len(nv)} nhan su: {len(pl01)} PL01 (co BHXH), {len(pl02)} PL02 (khong BHXH)")

    print("Xuat XML...")
    root = build_xml(nv, year)
    indent_xml(root)
    tree = ET.ElementTree(root)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
        f.write('<!-- Quyet toan thue TNCN | Luat 109/2025/QH15 | tncn-to-xml skill -->\n')
        ET.ElementTree(root).write(f, encoding="unicode", xml_declaration=False)

    # ── Thong ke ─────────────────────────────────────────────────────────────
    def fmt(n): return f"{int(n):,}".replace(",", ".")
    t_tn = sum(d["tn"] for d in nv.values())
    t_kt = sum(d["thue_kt"] for d in nv.values())
    t_pt = sum(
        calc_thue(max(0, d["tn"] - d["bhxh"] - d["cd"] - d["gtgc"]))
        for d in nv.values()
    )

    print()
    print("=" * 55)
    print("  XUAT XML HOAN THANH")
    print("=" * 55)
    print(f"  File      : {output_path}")
    print(f"  Nam       : {year}")
    print(f"  Nhan su   : {len(nv)} nguoi (PL01: {len(pl01)}, PL02: {len(pl02)})")
    print(f"  Tong TN   : {fmt(t_tn)} d")
    print(f"  Thue KT   : {fmt(t_kt)} d")
    print(f"  Thue PNop : {fmt(t_pt)} d")
    print(f"  Chenh lech: {fmt(t_pt - t_kt)} d")
    print("=" * 55)
    print()
    print("  Preview XML (15 dong dau):")
    print("  " + "-" * 50)
    with open(output_path, encoding="utf-8") as f:
        for i, line in enumerate(f):
            if i >= 15:
                print("  ...")
                break
            print("  " + line.rstrip())
    print()


if __name__ == "__main__":
    main()
