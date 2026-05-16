"""
Skill 2 - tncn-sheets-sync
Keo du lieu tu Google Sheets public, luu thanh file Excel thô
(bang_luong_from_sheets.xlsx) de Skill 1 va Skill 3 su dung.

Yeu cau: sheet phai duoc share "Anyone with the link - Viewer".
Chay: python .claude/skills/tncn-sheets-sync/scripts/fetch_and_update.py
"""
import sys, io, csv, json, urllib.request
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

CONFIG_PATH = Path(".claude/skills/tncn-sheets-sync/config.json")


def load_config():
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


def fetch_csv(spreadsheet_id, gid):
    url = (f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
           f"/export?format=csv&gid={gid}")
    try:
        with urllib.request.urlopen(url) as resp:
            if resp.status != 200:
                print(f"  ERROR: HTTP {resp.status}"); return []
            content = resp.read().decode("utf-8-sig")
    except Exception as e:
        print(f"  ERROR: {e}")
        print("  Kiem tra: sheet da duoc share 'Anyone with the link'?")
        return []
    rows = list(csv.reader(content.splitlines()))
    return rows


def write_sheet(ws, rows, sheet_name):
    if not rows:
        return 0

    # Header row style
    hdr_font  = Font(name="Arial", size=10, bold=True)
    hdr_fill  = PatternFill("solid", fgColor="BDD7EE")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)

    data_rows = 0
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font      = hdr_font
                cell.fill      = hdr_fill
                cell.alignment = hdr_align
            else:
                cell.font = data_font
                # Thu so thanh so nguyen neu co the
                try:
                    num = float(val.replace(",", "").replace(".", "")) if val else None
                    if num is not None and str(int(num)) == val.replace(",","").replace(".",""):
                        cell.value = int(num)
                except Exception:
                    pass
        if r_idx > 1:
            data_rows += 1

    # Do rong cot tu dong
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    return data_rows


if __name__ == "__main__":
    cfg = load_config()
    sid      = cfg["spreadsheet_id"]
    out_dir  = Path(cfg["output_dir"])
    out_file = out_dir / cfg["output_file"]
    out_dir.mkdir(exist_ok=True)

    print(f"Spreadsheet ID : {sid}")
    print(f"Output         : {out_file}")
    print()

    # Fetch 2 sheets
    print(f"Dang tai sheet '{cfg['sheet1_name']}' (gid={cfg['sheet1_gid']})...")
    rows1 = fetch_csv(sid, cfg["sheet1_gid"])
    print(f"  -> {max(len(rows1)-1,0)} dong du lieu")

    print(f"Dang tai sheet '{cfg['sheet2_name']}' (gid={cfg['sheet2_gid']})...")
    rows2 = fetch_csv(sid, cfg["sheet2_gid"])
    print(f"  -> {max(len(rows2)-1,0)} dong du lieu")

    if not rows1 and not rows2:
        print("\nERROR: Khong tai duoc du lieu. Dung lai."); sys.exit(1)

    # Ghi vao xlsx
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # xoa sheet mac dinh

    ws1 = wb.create_sheet(cfg["sheet1_name"])
    n1  = write_sheet(ws1, rows1, cfg["sheet1_name"])

    ws2 = wb.create_sheet(cfg["sheet2_name"])
    n2  = write_sheet(ws2, rows2, cfg["sheet2_name"])

    wb.save(str(out_file))

    print()
    print(f"Da luu: {out_file}")
    print(f"  Sheet '{cfg['sheet1_name']}': {n1} dong")
    print(f"  Sheet '{cfg['sheet2_name']}': {n2} dong")
    print(f"  Tong: {n1+n2} dong")
    print()
    print("Buoc tiep theo:")
    print("  /tncn-generator  ->  doc file nay, tao To khai quyet toan")
    print("  /tncn-review     ->  doc file nay, tao bao cao tong hop nhanh")
