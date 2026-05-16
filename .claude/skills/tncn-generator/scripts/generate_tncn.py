"""
Skill 1 - tncn-generator
Doc file Excel nhieu sheet bang luong, tinh thue TNCN theo Luat 109/2025/QH15,
dien vao mau chinh thuc 05/QTT-TNCN va xuat file ket qua.

Cach dung:
    python .claude/skills/tncn-generator/scripts/generate_tncn.py <bang_luong.xlsx>
"""
import sys, io
from datetime import datetime
from collections import defaultdict
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

try:
    import xlrd
    from xlutils.copy import copy as xl_copy
    import xlwt
except ImportError:
    print("ERROR: pip install xlrd xlutils xlwt")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ── Thue Luat 109/2025/QH15 ───────────────────────────────────────────────────
GIAM_TRU_BAN_THAN  = 15_500_000
GIAM_TRU_PHU_THUOC = 6_200_000
TEMPLATE  = Path("05_QTT_TNCN_TT80_2025.xls")
OUTPUT_DIR = Path("output")

# ── Mapping cot chinh xac theo mau goc ───────────────────────────────────────
# 05-1_BK-TNCN
PL01 = {
    "stt": 1, "ho_ten": 3, "mst": 4, "cccd": 7,
    "tnct": 12, "gtgc": 22, "bhxh": 26,
    "tn_tinh_thue": 30,
    "thue_kt": 32,   # [22] Tong so thue da khau tru
    "thue_pt": 36,   # [24] Thue phai nop
    "thue_thua": 38, # [25] Thue nop thua
    "thue_them": 40, # [26] Thue con phai nop
}
# 05-2_BK-TNCN
PL02 = {
    "stt": 1, "ho_ten": 3, "mst": 4, "cccd": 7,
    "tnct": 10,      # [11] Tong TNCT
    "thue_kt": 19,   # [15] So thue TNCN da khau tru - Tong so
}

def parse_vnd(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return int(round(v))
    try: return int(str(v).strip().replace(".","").replace(",","").replace(" ","") or 0)
    except: return 0

def calc_thue(tnct):
    if tnct <= 0: return 0
    t = min(tnct, 10_000_000) * 0.05
    if tnct > 10_000_000:  t += (min(tnct, 30_000_000)  - 10_000_000)  * 0.10
    if tnct > 30_000_000:  t += (min(tnct, 60_000_000)  - 30_000_000)  * 0.20
    if tnct > 60_000_000:  t += (min(tnct, 100_000_000) - 60_000_000)  * 0.30
    if tnct > 100_000_000: t += (tnct - 100_000_000) * 0.35
    return int(t)

# ── Styles xlwt ──────────────────────────────────────────────────────────────
def make_style(bg_color=None, bold=False, num_fmt=False, align="left"):
    style = xlwt.XFStyle()
    fnt = xlwt.Font()
    fnt.name  = "Arial"
    fnt.height = 200  # 10pt
    fnt.bold  = bold
    style.font = fnt

    if num_fmt:
        style.num_format_str = '#,##0'

    al = xlwt.Alignment()
    al.wrap = 1
    al.horz = xlwt.Alignment.HORZ_RIGHT if num_fmt else (
        xlwt.Alignment.HORZ_CENTER if align == "center" else xlwt.Alignment.HORZ_LEFT)
    style.alignment = al

    if bg_color is not None:
        pat = xlwt.Pattern()
        pat.pattern = xlwt.Pattern.SOLID_PATTERN
        pat.pattern_fore_colour = bg_color
        style.pattern = pat

    return style

# xlwt color indices (safe palette)
COLOR_HEADER = 0x1F   # pale blue
COLOR_TOTAL  = 0x16   # grey25
COLOR_NONE   = None

S_DATA_TEXT  = make_style()
S_DATA_NUM   = make_style(num_fmt=True)
S_HEADER     = make_style(bg_color=COLOR_HEADER, bold=True, align="center")
S_TOTAL_TEXT = make_style(bg_color=COLOR_TOTAL, bold=True)
S_TOTAL_NUM  = make_style(bg_color=COLOR_TOTAL, bold=True, num_fmt=True)

def w(ws, r, c, v, style=None):
    if style is None:
        style = S_DATA_NUM if isinstance(v, (int, float)) and v != "" else S_DATA_TEXT
    ws.write(r, c, v, style)

# ── Doc tat ca sheet tu file dau vao ─────────────────────────────────────────
def read_input(filepath):
    ext = Path(filepath).suffix.lower()
    all_rows = []
    INCOME_KEYS = {"họ tên","họ và tên","tổng thu nhập","thuế tncn","ho ten","tong thu nhap"}

    def process_sheet(rows_iter, sheet_name):
        rows = list(rows_iter)
        header_idx = None
        for i, row in enumerate(rows):
            cells = " ".join(str(c).lower().strip() if c else "" for c in row)
            if sum(1 for k in INCOME_KEYS if k in cells) >= 2:
                header_idx = i; break
        if header_idx is None: return 0
        headers = [str(c).strip() if c else "" for c in rows[header_idx]]
        count = 0
        for row in rows[header_idx + 1:]:
            if not any(c for c in row if c not in (None, "", 0)): continue
            rec = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
            name = str(next((rec[h] for h in rec if any(k in h.lower() for k in
                   ["họ tên","họ và tên","ho ten"])), "") or "").strip()
            if not name: continue
            rec["_sheet"] = sheet_name
            all_rows.append(rec); count += 1
        return count

    if ext == ".csv":
        import csv as _csv
        with open(filepath, encoding="utf-8-sig") as f:
            rows = list(_csv.reader(f))
        n = process_sheet(iter(rows), Path(filepath).stem)
        if n: print(f"  CSV '{Path(filepath).name}': {n} dong")
    elif ext == ".xls":
        wb = xlrd.open_workbook(filepath)
        for sname in wb.sheet_names():
            ws = wb.sheet_by_name(sname)
            rows_iter = ([ws.cell_value(r,c) for c in range(ws.ncols)] for r in range(ws.nrows))
            n = process_sheet(rows_iter, sname)
            if n: print(f"  Sheet '{sname}': {n} dong")
    elif ext in (".xlsx",".xlsm") and openpyxl:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows_iter = ([c.value for c in row] for row in ws.iter_rows())
            n = process_sheet(rows_iter, sname)
            if n: print(f"  Sheet '{sname}': {n} dong")
    else:
        print(f"ERROR: Dinh dang khong ho tro: {ext}")
        sys.exit(1)
    return all_rows

# ── Tong hop theo nhan vien ───────────────────────────────────────────────────
def aggregate(all_rows):
    def gv(rec, *keys):
        for k in keys:
            for hk in rec:
                if k.lower() in hk.lower():
                    v = parse_vnd(rec[hk])
                    if v: return v
        return 0

    def gs(rec, *keys):
        for k in keys:
            for hk in rec:
                if k.lower() in hk.lower():
                    v = str(rec[hk] or "").strip()
                    if v and v not in ("0","None"): return v
        return ""

    # Buoc 1: gom theo (ten, ky tra luong) — tranh cong GTGC nhieu lan tren cung 1 ky
    periods = defaultdict(lambda: {
        "tn":0,"thue_kt":0,"bhxh":0,"cd":0,"gtgc":0,
        "cccd":"","mst":"","has_bhxh":False
    })

    for rec in all_rows:
        name = str(next((rec[h] for h in rec if any(k in h.lower()
               for k in ["họ tên","họ và tên","ho ten"])), "") or "").strip()
        if not name: continue

        period = gs(rec, "tháng trả lương","thang tra luong","tháng","thang") or "default"
        key = (name, period)

        bhxh = gv(rec, "BHXH khấu trừ","bhxh khau tru","bhxh")
        gtgc = gv(rec, "Giảm trừ gia cảnh","giam tru gia canh","giam tru")

        periods[key]["tn"]      += gv(rec,"Tổng thu nhập","tong thu nhap","Số tiền","so tien")
        periods[key]["thue_kt"] += gv(rec,"Thuế TNCN đã khấu trừ","thue tncn da khau tru","Thuế TNCN","thue tncn")
        periods[key]["bhxh"]    += bhxh
        periods[key]["cd"]      += gv(rec,"Công đoàn","cong doan")
        periods[key]["gtgc"]    += gtgc   # cong thang gia tri co trong cot, khong ap cong thuc
        if bhxh > 0: periods[key]["has_bhxh"] = True
        if not periods[key]["cccd"]: periods[key]["cccd"] = gs(rec,"cccd","cmnd","so cccd","so dinh danh")
        if not periods[key]["mst"]:  periods[key]["mst"]  = gs(rec,"ma so thue","mst")

    # Buoc 2: gom theo ten nguoi (cong don qua cac ky)
    nv = defaultdict(lambda: {
        "tn":0,"thue_kt":0,"bhxh":0,"cd":0,"gtgc":0,
        "cccd":"","mst":"","has_bhxh":False
    })

    for (name, period), d in periods.items():
        nv[name]["tn"]      += d["tn"]
        nv[name]["thue_kt"] += d["thue_kt"]
        nv[name]["bhxh"]    += d["bhxh"]
        nv[name]["cd"]      += d["cd"]
        nv[name]["gtgc"]    += d["gtgc"]   # cong thang, khong them mac dinh
        if d["has_bhxh"]: nv[name]["has_bhxh"] = True
        if not nv[name]["cccd"]: nv[name]["cccd"] = d["cccd"]
        if not nv[name]["mst"]:  nv[name]["mst"]  = d["mst"]

    return nv

# ── Ghi vao mau .xls ─────────────────────────────────────────────────────────
def write_output(nv, year, output_path):
    rb = xlrd.open_workbook(str(TEMPLATE), formatting_info=True)
    wb = xl_copy(rb)

    pl01 = {n:d for n,d in nv.items() if d["has_bhxh"]}
    pl02 = {n:d for n,d in nv.items() if not d["has_bhxh"]}

    t_tn = sum(d["tn"] for d in nv.values())
    t_kt = sum(d["thue_kt"] for d in nv.values())
    t_pt = sum(calc_thue(max(0,d["tn"]-d["bhxh"]-d["cd"]-d["gtgc"])) for d in nv.values())

    snames = rb.sheet_names()
    def get_ws(name): return wb.get_sheet(snames.index(name))

    # ── To khai chinh ──────────────────────────────────────────────────────
    tk = get_ws("Tờ khai")
    tk.write(2, 10, f"Kỳ tính thuế:  Năm {year}", S_DATA_TEXT)
    tk.write(47, 20, len(nv),    S_DATA_NUM)   # [16] tong NLD
    tk.write(48, 20, len(pl01),  S_DATA_NUM)   # [17] co HĐ
    tk.write(49, 20, len(nv),    S_DATA_NUM)   # [18] da khau tru
    tk.write(50, 20, len(nv),    S_DATA_NUM)   # [19] cu tru
    tk.write(51, 20, 0,          S_DATA_NUM)   # [20] khong cu tru
    tk.write(59, 20, t_tn,       S_DATA_NUM)   # [23] tong TNCT

    # ── Phu luc 01 ────────────────────────────────────────────────────────
    ws1 = get_ws("05-1_BK-TNCN")
    ws1.write(4, 12, f"Kỳ tính thuế:  Năm {year}", S_DATA_TEXT)

    DATA_ROW = 23
    t1_tn=t1_bhxh=t1_gtgc=t1_tnct=t1_kt=t1_pt = 0

    for i, (name, d) in enumerate(sorted(pl01.items())):
        r = DATA_ROW + i
        tnct = max(0, d["tn"] - d["bhxh"] - d["cd"] - d["gtgc"])
        pt   = calc_thue(tnct)
        cl   = pt - d["thue_kt"]

        ws1.write(r, PL01["stt"],          i+1,          S_DATA_NUM)
        ws1.write(r, PL01["ho_ten"],       name,         S_DATA_TEXT)
        ws1.write(r, PL01["mst"],          d["mst"],     S_DATA_TEXT)
        ws1.write(r, PL01["cccd"],         d["cccd"],    S_DATA_TEXT)
        ws1.write(r, PL01["tnct"],         d["tn"],      S_DATA_NUM)
        ws1.write(r, PL01["bhxh"],         d["bhxh"],    S_DATA_NUM)
        ws1.write(r, PL01["gtgc"],         d["gtgc"],    S_DATA_NUM)
        ws1.write(r, PL01["tn_tinh_thue"], tnct,         S_DATA_NUM)
        ws1.write(r, PL01["thue_kt"],      d["thue_kt"], S_DATA_NUM)
        ws1.write(r, PL01["thue_pt"],      pt,           S_DATA_NUM)
        if cl < 0: ws1.write(r, PL01["thue_thua"], abs(cl), S_DATA_NUM)
        if cl > 0: ws1.write(r, PL01["thue_them"], cl,      S_DATA_NUM)

        t1_tn+=d["tn"]; t1_bhxh+=d["bhxh"]; t1_gtgc+=d["gtgc"]
        t1_tnct+=tnct;  t1_kt+=d["thue_kt"]; t1_pt+=pt

    # Dong tong PL01
    tr1 = DATA_ROW + len(pl01)
    ws1.write(tr1, PL01["stt"],          "Tổng",  S_TOTAL_TEXT)
    ws1.write(tr1, PL01["tnct"],         t1_tn,   S_TOTAL_NUM)
    ws1.write(tr1, PL01["bhxh"],         t1_bhxh, S_TOTAL_NUM)
    ws1.write(tr1, PL01["gtgc"],         t1_gtgc, S_TOTAL_NUM)
    ws1.write(tr1, PL01["tn_tinh_thue"], t1_tnct, S_TOTAL_NUM)
    ws1.write(tr1, PL01["thue_kt"],      t1_kt,   S_TOTAL_NUM)
    ws1.write(tr1, PL01["thue_pt"],      t1_pt,   S_TOTAL_NUM)

    # ── Phu luc 02 ────────────────────────────────────────────────────────
    ws2 = get_ws("05-2_BK-TNCN")
    ws2.write(4, 12, f"Kỳ tính thuế:  Năm {year}", S_DATA_TEXT)

    t2_tn=t2_kt = 0

    for i, (name, d) in enumerate(sorted(pl02.items())):
        r = DATA_ROW + i
        ws2.write(r, PL02["stt"],     i+1,          S_DATA_NUM)
        ws2.write(r, PL02["ho_ten"],  name,         S_DATA_TEXT)
        ws2.write(r, PL02["mst"],     d["mst"],     S_DATA_TEXT)
        ws2.write(r, PL02["cccd"],    d["cccd"],    S_DATA_TEXT)
        ws2.write(r, PL02["tnct"],    d["tn"],      S_DATA_NUM)
        ws2.write(r, PL02["thue_kt"], d["thue_kt"], S_DATA_NUM)
        t2_tn+=d["tn"]; t2_kt+=d["thue_kt"]

    # Dong tong PL02
    tr2 = DATA_ROW + len(pl02)
    ws2.write(tr2, PL02["stt"],     "Tổng", S_TOTAL_TEXT)
    ws2.write(tr2, PL02["tnct"],    t2_tn,  S_TOTAL_NUM)
    ws2.write(tr2, PL02["thue_kt"], t2_kt,  S_TOTAL_NUM)

    wb.save(str(output_path))

# ── Chat history ─────────────────────────────────────────────────────────────
def write_chat_history(nv, input_file, output_path, year):
    pl01 = {n:d for n,d in nv.items() if d["has_bhxh"]}
    pl02 = {n:d for n,d in nv.items() if not d["has_bhxh"]}
    t_tn = sum(d["tn"] for d in nv.values())
    t_kt = sum(d["thue_kt"] for d in nv.values())
    t_pt = sum(calc_thue(max(0,d["tn"]-d["bhxh"]-d["cd"]-d["gtgc"])) for d in nv.values())
    def fmt(n): return f"{int(n):,}".replace(",",".")

    lines = [
        "=" * 60,
        "  CLAUDE CODE — CHAT HISTORY",
        "  Skill: tncn-generator",
        "=" * 60,
        f"Thoi gian    : {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        f"File dau vao : {input_file}",
        f"Luat thue    : Luat 109/2025/QH15",
        f"Nam tinh thue: {year}",
        "",
        "KET QUA XU LY:",
        f"  Phu luc 01 (co BHXH)    : {len(pl01):>5} nguoi",
        f"  Phu luc 02 (khong BHXH) : {len(pl02):>5} nguoi",
        f"  Tong nhan su            : {len(nv):>5} nguoi",
        "",
        "CHI SO TAI CHINH:",
        f"  Tong thu nhap da tra     : {fmt(t_tn):>22} VND",
        f"  Thue TNCN da khau tru   : {fmt(t_kt):>22} VND",
        f"  Thue TNCN phai nop (L109): {fmt(t_pt):>22} VND",
        f"  Chenh lech              : {fmt(t_pt-t_kt):>22} VND",
        "",
        "FILE DA TAO:",
        f"  {output_path}",
        "    Sheet To khai chinh : 05/QTT-TNCN",
        f"    Sheet Phu luc 01   : 05-1/BK-TNCN ({len(pl01)} nguoi co BHXH)",
        f"    Sheet Phu luc 02   : 05-2/BK-TNCN ({len(pl02)} nguoi khong BHXH)",
        "    Sheet Phu luc 03   : 05-3/BK-TNCN (de trong)",
        "=" * 60,
    ]
    hist = OUTPUT_DIR / f"chat_history_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.txt"
    with open(hist, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  Chat history: {hist}")

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Cach dung: python generate_tncn.py <file_bang_luong>")
        print("Vi du   : python generate_tncn.py 'Form bang luong 2026.xlsx'")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    if not input_path.exists():
        print(f"ERROR: Khong tim thay '{input_path}'"); sys.exit(1)
    if not TEMPLATE.exists():
        print(f"ERROR: Khong tim thay mau '{TEMPLATE}'"); sys.exit(1)

    year = datetime.now().year
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_path = OUTPUT_DIR / f"QuyetToanTNCN_{year}.xls"

    print(f"Doc du lieu: {input_path}")
    all_rows = read_input(str(input_path))
    print(f"Tong {len(all_rows)} dong")

    print("Tong hop...")
    nv = aggregate(all_rows)
    pl01 = {n:d for n,d in nv.items() if d["has_bhxh"]}
    pl02 = {n:d for n,d in nv.items() if not d["has_bhxh"]}
    print(f"  {len(nv)} nhan vien: {len(pl01)} PL01, {len(pl02)} PL02")

    print("Dien vao mau chinh thuc...")
    write_output(nv, year, output_path)
    write_chat_history(nv, input_path, output_path, year)

    print(f"\nHoan thanh! Output: {output_path}")
